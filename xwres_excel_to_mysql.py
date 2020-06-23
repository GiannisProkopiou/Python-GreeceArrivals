import openpyxl
import xlrd
import MySQLdb
import mysql.connector

def xwres_excel_to_mysql():

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2011_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2011_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2011_a( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if(number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2011_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2011_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2011_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2011_b( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2011_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2011_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2011_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2011_c( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2011_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2011_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2011_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2011_d( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2011_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2012_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2012_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2012_a( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2012_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2012_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2012_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2012_b( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2012_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2012_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠΤ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2012_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2012_c( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2012_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2012_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2012_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2012_d( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2012_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2013_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2013_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2013_a( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2013_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2013_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2013_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2013_b( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2013_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2013_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠΤ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2013_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2013_c( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2013_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2013_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2013_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2013_d( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2013_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2014_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2014_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2014_a( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2014_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2014_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2014_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2014_b( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2014_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2014_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠΤ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2014_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2014_c( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2014_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2014_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2014_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2014_d( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2014_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2015_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2015_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2015_a( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2015_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2015_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2015_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2015_b( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2015_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2015_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠΤ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2015_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2015_c( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2015_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/xwra_2015_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    #data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚΕΜ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS xwra_2015_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, year_before varchar(255) , year_now varchar(255), year_difference varchar(255), year_analogy varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO xwra_2015_d( number , nation, year_before, year_now, year_difference, year_analogy) VALUES (%s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(78, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        year_before = data_sheet.cell(r, 2).value
        year_now = data_sheet.cell(r, 3).value
        year_difference = data_sheet.cell(r, 4).value
        year_analogy = data_sheet.cell(r, 5).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, year_before, year_now, year_difference, year_analogy)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, year_before, year_now, year_difference, year_analogy FROM xwra_2015_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_xwra_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()