import openpyxl
import xlrd
import MySQLdb
import mysql.connector

def meso_excel_to_mysql():

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2011_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2011_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2011_a( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2011_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2011_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2011_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2011_b( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total )

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2011_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2011_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2011_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2011_c( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2011_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2011_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2011_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2011_d( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2011_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2012_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2012_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2012_a( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2012_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2012_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2012_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2012_b( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2012_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2012_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠΤ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2012_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2012_c( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2012_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2012_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2012_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2012_d( number , nation, plane, train, ship, car, total ) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2012_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2013_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2013_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2013_a( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2013_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2013_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2013_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2013_b( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2013_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2013_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2013_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2013_c( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2013_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2013_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2013_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2013_d( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2013_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2014_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2014_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2014_a( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car,total FROM meso_2014_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2014_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΙΟΥΝ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2014_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2014_b( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2014_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2014_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠΤ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2014_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2014_c( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car,total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2014_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2014_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2014_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2014_d( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2014_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2015_a.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2015_a(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2015_a( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2015_a")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2015_b.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΜΑΡ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2015_b(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2015_b( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2015_b")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2015_c.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΣΕΠΤ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2015_c(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2015_c( number , nation, plane, train, ship, car, total) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2015_c")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()

    # Open the workbook and define the worksheet
    excel_file_path: str = 'D:/Python_Project_Compilers/meso_2015_d.xls'
    work_sheet = xlrd.open_workbook(excel_file_path)
    # data_sheet = work_sheet.sheet_by_index(0)
    data_sheet = work_sheet.sheet_by_name('ΔΕΚΕΜ')

    # Establish a MySQL connection
    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # First create the data base
    mycursor.execute("CREATE  DATABASE IF NOT EXISTS TOURISM_DATABASE")

    try:
        database = mysql.connector.connect(host="localhost", user="root", passwd="", database="TOURISM_DATABASE")
    except:
        print('Database connection failed!')

    # Get the cursor, which is used to traverse the database, line by line
    mycursor = database.cursor()

    # Create Table
    mycursor.execute(
        "CREATE TABLE IF NOT EXISTS meso_2015_d(number varchar(255),nation NVARCHAR(255) PRIMARY KEY, plane varchar(255) , train varchar(255), ship varchar(255), car varchar(255), total varchar(255))")

    # Create the INSERT INTO sql query
    query = """INSERT INTO meso_2015_d( number , nation, plane, train, ship, car, total ) VALUES (%s, %s, %s, %s, %s, %s, %s)"""

    # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
    for r in range(76, data_sheet.nrows):
        number = data_sheet.cell(r, 0).value
        nation = data_sheet.cell(r, 1).value
        plane = data_sheet.cell(r, 2).value
        train = data_sheet.cell(r, 3).value
        ship = data_sheet.cell(r, 4).value
        car = data_sheet.cell(r, 5).value
        total = data_sheet.cell(r, 6).value

        try:
            if (number == ''):
                continue

            # Assign values from each row
            values = (number, nation, plane, train, ship, car, total)

            # Execute sql Query
            mycursor.execute(query, values)
        except:
            continue

    # Select Given
    mycursor.execute("SELECT number, nation, plane, train, ship, car, total FROM meso_2015_d")
    select_result = mycursor.fetchall()

    for x in select_result:
        print(x)

    # Fetch only first row
    # mycursor.execute("SELECT * FROM 2011_meso_a")
    # result_fetch_first_row = mycursor.fetchone()
    print(data_sheet.nrows)
    # print(result_fetch_first_row)
    # Close the cursor
    mycursor.close()

    # Commit the transaction
    database.commit()

    # Print Insterted
    # print(mycursor.rowcount, "was inserted.")

    # Close the database connection
    database.close()